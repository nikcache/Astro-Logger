from graphics import *
from txt import *
from buttonclass import *
import datetime
from tkinter import filedialog
import tkinter as tk
import astropy.io
from astropy.io import fits
from cbox import *
import openpyxl
from openpyxl import *

def EntryWindow():

    #Processes
    win = GraphWin("Log Window", 600, 800)

    #Logger Name
    txt("Full Name:", 50, 25, win, 12, "black", "normal")
    nameIn = Entry(Point(230, 25), 30)
    nameIn.draw(win)

    #Date Grabber
    today = datetime.datetime.today()

    #Date 
    txt("Date:", 400, 25, win, 12, "black", "normal")

    #Year
    txt("YYYY", 440, 50, win, 12, "black", "normal")
    YearIn = Entry(Point(440, 25), 4)
    YearIn.setText(today.year)
    YearIn.draw(win)

    #Month
    txt("MM", 480, 50, win, 12, "black", "normal")
    MonthIn = Entry(Point(480, 25), 3)
    MonthIn.setText(today.strftime("%m"))
    MonthIn.draw(win)

    #Day
    txt("DD", 515, 50, win, 12, "black", "normal")
    DayIn = Entry(Point(515, 25), 3)
    DayIn.setText(today.strftime("%d"))
    DayIn.draw(win)

    #Clear Mode Button
    clearButton = Button(win, Point(565,25), 40, 25, "Clear")

    #QuickAccess Buttons
    NikButton = Button(win, Point(120,55), 40, 25, "Nik")
    NikeshButton = Button(win, Point(175,55), 60, 25, "Nikesh")
    DocButton = Button(win, Point(260,55), 100, 25, "Doc Brown")

    #Divider
    divide1 = Rectangle(Point(25, 75), Point(575, 80))
    divide1.setFill("black")
    divide1.draw(win)

    #First Picture Selector
    txt("First Picture File:", 75, 100, win, 12, "black", "normal")
    selectButton1 = Button(win, Point(250, 100), 175, 25, "Select First Picture")

    #Second Picture Selector
    selectButton2 = Button(win, Point(450, 100), 175, 25, "Select Second Picture")
    
    #Header Info
    #Observed Object
    txt("Object Observed: ", 100, 130, win, 15, "black", "bold")
    Object = Text(Point(300, 130),"--")
    Object.draw(win)

    #Telescope Used
    txt("Telescope Used:  ", 100, 160, win, 12, "black", "bold")
    telescope = Text(Point(300, 160),"--")
    telescope.draw(win)

    #Temp of CCD
    txt("Temperature CCD: ", 100, 190, win, 12, "black", "bold")
    temp = Text(Point(300, 190),"--")
    temp.draw(win)

    #Bands boxes
    txt("Bands Used", 512, 140, win, 12, "black", "bold")
    
    bBandBox = checkbox(475, 170, win, "B")
    bBandVal = False

    vBandBox = checkbox(500, 170, win, "V")
    vBandVal = False

    rBandBox = checkbox(525, 170, win, "R")
    rBandVal = False

    iBandBox = checkbox(550, 170, win, "I")
    iBandVal = False

    #Divider
    divide2 = Rectangle(Point(25, 210), Point(575, 215))
    divide2.setFill("black")
    divide2.draw(win)

    #First Picture info
    #Title
    txt("First Picture Info", 90, 230, win, 15, "black", "bold")
    txt("(Pick first picture taken that day)", 285, 230, win, 12, "black", "normal")

    #Date and time UTC
    txt("Date and time: ", 83, 260, win, 12, "black", "bold")
    dateUTC = Text(Point(300, 260),"--")
    dateUTC.draw(win)

    #Julian Date
    txt("Julian Date: ", 83, 290, win, 12, "black", "bold")
    JulianDate = Text(Point(300, 290),"--")
    JulianDate.draw(win)

    #Airmass
    txt("Airmass:     ", 83, 320, win, 12, "black", "bold")
    Airmass = Text(Point(300, 320),"--")
    Airmass.draw(win)


    #Divider
    divide3 = Rectangle(Point(25, 340), Point(575, 345))
    divide3.setFill("black")
    divide3.draw(win)


    #Last Picture info
    #Title
    txt("Last Picture Info", 90, 365, win, 15, "black", "bold")
    txt("(Pick last picture taken that day)", 285, 365, win, 12, "black", "normal")

    #Date and time UTC
    txt("Date and time: ", 83, 405, win, 12, "black", "bold")
    dateUTC2 = Text(Point(300, 405),"--")
    dateUTC2.draw(win)

    #Julian Date
    txt("Julian Date: ", 83, 435, win, 12, "black", "bold")
    JulianDate2 = Text(Point(300, 435),"--")
    JulianDate2.draw(win)

    #Airmass
    txt("Airmass:     ", 83, 465, win, 12, "black", "bold")
    Airmass2 = Text(Point(300, 465),"--")
    Airmass2.draw(win)
    

    #Divider
    divide4 = Rectangle(Point(25, 485), Point(575, 490))
    divide4.setFill("black")
    divide4.draw(win)

    #Checkboxes

    txt("Please check all that apply: ", 150, 515, win, 15, "black", "bold")

    txt("Calibrated using MaxIm DL", 145, 550, win, 12, "black", "normal")
    maximCal = checkbox(25, 550, win, "")
    maximCal.setTrue()
    maximCalVal = maximCal.getVal()

    txt("Calibrated using Pyraf or IRAF", 155, 580, win, 12, "black", "normal")
    otherCal = checkbox(25, 580, win, "")
    otherCalVal = otherCal.getVal()

    txt("Subtracted Master Bias, Dark and Flats?", 235, 610, win, 12, "black", "normal")
    biasBox = checkbox(25, 610, win, "B")
    biasBox.setTrue()
    biasVal = biasBox.getVal()

    darkBox = checkbox(50, 610, win, "M")
    darkBox.setTrue()
    darkVal = darkBox.getVal()

    flatBox = checkbox(75, 610, win, "F")
    flatBox.setTrue()
    flatVal = flatBox.getVal()

    txt("Did you do all those steps one by one?", 185, 640, win, 12, "black", "normal")
    stepBox = checkbox(25, 640, win, "")
    stepVal = stepBox.getVal()

    txt("Did you use auto-calibration for the files?", 190, 670, win, 12, "black", "normal")
    autoBox = checkbox(25, 670, win, "")
    autoBox.setTrue()
    autoVal = autoBox.getVal()

    #Notes
    txt("Notes:", 43, 710, win, 15, "black", "bold")
    notesEntry = Entry(Point(330, 710), 55)
    notesEntry.draw(win)

    #Save Button
    saveButton = Button(win, Point(300, 745), 50, 25, "Save")

    #Help Button
    helpButton = Button(win, Point(550, 745), 50, 25, "Help")

    #Photometry Button
    photoButton = Button(win, Point(465, 745), 100, 25, "Photometry")

    #Quit Button
    quitButton = Button(win, Point(300, 775), 50, 25, "Quit")

    #Notification text
    txt1 = Text(Point(150, 745), "")
    txt1.draw(win)

    #Instance List
    photoDetails = ["", "", "", "", ""]

    
    pt = win.getMouse()
    while not quitButton.isClicked(pt):

        if (clearButton.isClicked(pt)):
            nameIn.setText("")
            YearIn.setText("")
            MonthIn.setText("")
            DayIn.setText("")

        if (NikButton.isClicked(pt)):
            nameIn.setText("Nik Korzoun")

        if (NikeshButton.isClicked(pt)):
            nameIn.setText("Nikesh Ghimire")

        if (DocButton.isClicked(pt)):
            nameIn.setText("Leslie Brown")

        if (bBandBox.isClicked(pt)):
            bBandVal = bBandBox.getVal()

        if (vBandBox.isClicked(pt)):
            vBandVal = vBandBox.getVal()

        if (rBandBox.isClicked(pt)):
            rBandVal = rBandBox.getVal()

        if (iBandBox.isClicked(pt)):
            iBandVal = iBandBox.getVal()

        if (maximCal.isClicked(pt)):
            maximCalVal = maximCal.getVal()

        if (biasBox.isClicked(pt)):
            biasVal = biasBox.getVal()

        if (darkBox.isClicked(pt)):
            darkVal = darkBox.getVal()

        if (flatBox.isClicked(pt)):
            flatVal = flatBox.getVal()

        if (otherCal.isClicked(pt)):
            otherCalVal = otherCal.getVal()

        if (stepBox.isClicked(pt)):
            stepVal = stepBox.getVal()

        if (autoBox.isClicked(pt)):
            autoVal = autoBox.getVal()

        if (selectButton1.isClicked(pt)):
            try:
                root = tk.Tk()
                root.withdraw()
                file_path = filedialog.askopenfilename(initialdir = "/Users/Nikesh Ghimire/Documents/Astro",title\
                     = "Select file",filetypes = (("FITS files","*.fts *.fit"),("all files","*.*")))
                firstFile = fits.open(file_path)

                #Object Locater
                Object_name = firstFile[0].header["Object"]
                Object.setText(Object_name)

                #Telescope Locater
                TelescopeUsed = firstFile[0].header["TELESCOP"]
                telescope.setText(TelescopeUsed)

                #Temperature Locater
                TempCCD = str(firstFile[0].header["CCD-TEMP"])  + " °C"
                temp.setText(TempCCD)

                #Date Locater
                dateFirst = firstFile[0].header["DATE-OBS"] + " UTC"
                dateUTC.setText(dateFirst)

                #Julian Date Locater
                JulianFirst = firstFile[0].header["JD"]
                JulianDate.setText(JulianFirst)

                #Airmass Locater
                try:
                    AirmassFirst = firstFile[0].header["AIRMASS"]
                    Airmass.setText(AirmassFirst)
                    airmass1 = True
                except KeyError:
                    Airmass.setText("No Airmass data")
                    airmass1 = False
            except ValueError:
                txt1.setText("Empty Input")
                Object.setText("No file chosen")
                dateUTC.setText("No file chosen")
            except OSError:
                txt1.setText("File not supported")
                Object.setText("No file chosen")
                dateUTC.setText("No file chosen")

        if (selectButton2.isClicked(pt)):

            try:
                root2 = tk.Tk()
                root2.withdraw()
                file_path2 = filedialog.askopenfilename(initialdir = "/Users/Nikesh Ghimire/Documents/Astro",title\
                     = "Select file",filetypes = (("FITS files","*.fts *.fit"),("all files","*.*")))
                lastFile = fits.open(file_path2)

                #Date Locater
                dateLast = lastFile[0].header["DATE-OBS"] + " UTC"
                dateUTC2.setText(dateLast)

                #Julian Date Locater
                JulianLast = lastFile[0].header["JD"]
                JulianDate2.setText(JulianLast)

                #Airmass Locater
                try:
                    AirmassLast = lastFile[0].header["AIRMASS"]
                    Airmass2.setText(AirmassLast)
                    airmass2 = True
                except KeyError:
                    Airmass2.setText("No Airmass data")
                    airmass2 = False
            except ValueError:
                dateUTC2.setText("No file chosen")

        if (helpButton.isClicked(pt)):
            helpWin()

        if (photoButton.isClicked(pt)):
            photoDetails = photoWin(photoDetails)
            

    #Writing to Excel File

    #Loading file
        if (saveButton.isClicked(pt)):
            try:
                try:
                    wb = load_workbook("AstroLog2019.xlsx")
                except FileNotFoundError:
                    wb = Workbook()

                #Checking if source object already has logs
                if Object_name not in wb.sheetnames:
                    wb.create_sheet(Object_name)

                    #Making sheet active
                    ws = wb[Object_name]

                    #Creating Titles on sheet
                    ws["A1"] = "Log Date"
                    ws["B1"] = "Logger Name"
                    ws["C1"] = "Date of Observation"
                    ws["D1"] = "Temperature of CCD"
                    ws["E1"] = "Light Bands observed"
                    ws["F1"] = "Date and time of first picture"
                    ws["G1"] = "Julian date of first picture"
                    ws["H1"] = "Airmass of first picture"
                    ws["I1"] = "Date and time of last picture"
                    ws["J1"] = "Julian date of last picture"
                    ws["K1"] = "Airmass of last picture"
                    ws["L1"] = "Telescope Used"
                    ws["M1"] = "Calibrated using Maxim DL"
                    ws["N1"] = "Calibrated using Pyraf or IRAF"
                    ws["O1"] = "Subtracted Master Bias?"
                    ws["P1"] = "Subtracted Master Dark?"
                    ws["Q1"] = "Divided Master Flat(s)?"
                    ws["R1"] = "Steps done one by one?"
                    ws["S1"] = "Used auto-calibration?"
                    ws["T1"] = "Notes about Observation"
                    ws["U1"] = "URL of source (Photometry)"
                    ws["V1"] = "Number of reference stars used"
                    ws["W1"] = "Aperture diameter (pixels and arcseconds)"
                    ws["X1"] = "Sky annulus inner and outer radii or diameters"
                    ws["Y1"] = "Photometry Notes"

                else:
                    ws = wb[Object_name]

                #Extending existing Logs
                
                if (ws.cell(row = 1, column = 26).value == None):    
                    ws["Z1"] = "File Path"
                    
                #Adding Info
                rowE = ws.max_row+1
                ws.cell(row = rowE, column = 1).value = YearIn.getText() + "-" +\
                                                                MonthIn.getText() + "-"+\
                                                                DayIn.getText()
                ws.cell(row = rowE, column = 2).value = nameIn.getText()
                ws.cell(row = rowE, column = 3).value = dateFirst
                ws.cell(row = rowE, column = 4).value = TempCCD
                bandText = ""
                if (bBandVal):
                    bandText = bandText + "B"
                if (vBandVal):
                    bandText = bandText + "V"
                if (rBandVal):
                    bandText = bandText + "R"
                if (iBandVal):
                    bandText = bandText + "I"
                ws.cell(row = rowE, column = 5).value = bandText
                ws.cell(row = rowE, column = 6).value = dateFirst
                ws.cell(row = rowE, column = 7).value = JulianFirst
                if airmass1:
                    ws.cell(row = rowE, column = 8).value = AirmassFirst
                else:
                    ws.cell(row = rowE, column = 8).value = "No data"
                ws.cell(row = rowE, column = 9).value = dateLast
                ws.cell(row = rowE, column = 10).value = JulianLast
                if airmass2:
                    ws.cell(row = rowE, column = 11).value = AirmassLast
                else:
                    ws.cell(row = rowE, column = 11).value = "No data"
                ws.cell(row = rowE, column = 12).value = TelescopeUsed
                ws.cell(row = rowE, column = 13).value = maximCalVal
                ws.cell(row = rowE, column = 14).value = otherCalVal
                ws.cell(row = rowE, column = 15).value = biasVal
                ws.cell(row = rowE, column = 16).value = darkVal
                ws.cell(row = rowE, column = 17).value = flatVal
                ws.cell(row = rowE, column = 18).value = stepVal
                ws.cell(row = rowE, column = 19).value = autoVal
                ws.cell(row = rowE, column = 20).value = notesEntry.getText()
                ws.cell(row = rowE, column = 21).value = photoDetails[0]
                ws.cell(row = rowE, column = 22).value = photoDetails[1]
                ws.cell(row = rowE, column = 23).value = photoDetails[2]
                ws.cell(row = rowE, column = 24).value = photoDetails[3]
                ws.cell(row = rowE, column = 25).value = photoDetails[4]
                ws.cell(row = rowE, column = 26).value = file_path
                

                #Resetting Notes
                notesEntry.setText("")
                photoDetails[4] = ""
                
                #Closing and saving file
                txt1.setText("Saved, click to continue")
                wb.save("AstroLog2019.xlsx")
                wb.close()
            except UnboundLocalError:
                txt1.setText("Missing some values, try again")
        pt = win.getMouse()
        try:
            txt1.setText("")
        except UnboundLocalError:
            x = 1

    
    txt1.setText("Click to close")
    
    win.getMouse()    
    win.close()


#Help Window
def helpWin():
    hwin = GraphWin("Help Window", 600, 300)

    txt("Saving Conventions", 300, 25, hwin, 15, "black", "bold")
    txt("Raw images should be placed\n in a folder with name format 'YYYYMMDD_SOURCENAME'",\
        300, 75, hwin, 12, "black", "normal")
    txt("Calibrated images for that day's images\
        \nshould be placed inside 'CAL' folder inside raw folder",\
        300, 125, hwin, 12, "black", "normal")
    txt("If calibration files were specific for any day\n\
        place those files inside 'CAL FILES' inside raw folder",\
        300, 175, hwin, 12, "black", "normal")
    txt("The logger is completely open-source,\n\
        please email nghimir1@conncoll.edu for access and questions",\
        300, 225, hwin, 12, "black", "normal")
    txt("Click to close", 300, 275, hwin, 12, "black", "normal")

    hwin.getMouse()
    hwin.close()

def photoWin(inList):
    pwin = GraphWin("Photometry Log Window", 600, 400)

    txt("Photometry Log", 300, 25, pwin, 15, "black", "bold")
    txt("What is the URL of the site or the full bibliographic\nreference of the finder chart and reference\n star magnitudes and filters used to photometer this data?", 300, 75, pwin, 12, "black", "normal")
    txt("URL/Paper Reference:", 100, 125, pwin, 12, "black", "normal")

    urlIn = Entry(Point(380, 125), 42)
    urlIn.setText(inList[0])
    urlIn.draw(pwin)

    txt("How many reference stars were used?", 150, 175, pwin, 12, "black", "normal")

    refNumIn = Entry(Point(320, 175), 5)
    refNumIn.setText(inList[1])
    refNumIn.draw(pwin)

    txt("Aperture diameter in pixels and arcseconds:", 172, 225, pwin, 12, "black", "normal")

    apeIn = Entry(Point(380, 225), 10)
    apeIn.setText(inList[2])
    apeIn.draw(pwin)

    txt("Sky annulus inner and outer radii or diameters:", 180, 275, pwin, 12, "black", "normal")

    annIn = Entry(Point(375, 275), 5)
    annIn.setText(inList[3])
    annIn.draw(pwin)

    txt("Notes:", 43, 325, pwin, 12, "black", "bold")

    pNotesIn = Entry(Point(330, 325), 55)
    pNotesIn.setText(inList[4])
    pNotesIn.draw(pwin)

    submitButton = Button(pwin, Point(300, 375), 50, 25, "Save")

    pt = pwin.getMouse()

    while not submitButton.isClicked(pt):
        pt = pwin.getMouse()
    returnList = [urlIn.getText(), refNumIn.getText(), apeIn.getText(), annIn.getText(), pNotesIn.getText()]
    pwin.close()
    
    return returnList
     
EntryWindow()
